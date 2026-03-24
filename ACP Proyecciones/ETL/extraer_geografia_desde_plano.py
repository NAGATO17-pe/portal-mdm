"""
extraer_geografia_desde_plano.py
================================
Extrae geografia estructurada desde el plano de cultivo en Excel.

Busca codigos incrustados en el mapa como:
  - M6-T4-V8
  - S7M1T1V4

Salida:
  - geografia_plano_extraida.csv
  - seed_mdm_catalogo_geografia.sql

Uso:
  py extraer_geografia_desde_plano.py --archivo "ruta.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


PATRON_MTV = re.compile(r'^M(\d+)-T(\d+)-V(\d+)$', re.I)
PATRON_SMTV = re.compile(r'^S(\d+)M(\d+)T(\d+)V(\d+)$', re.I)


def _int_or_none(valor: str | None) -> int | None:
    if valor in (None, ''):
        return None
    return int(valor)


def extraer_registros(ruta_excel: Path,
                      fundo_default: str | None = None,
                      sector_default: str | None = None) -> list[dict]:
    wb = load_workbook(ruta_excel, data_only=True)
    registros: dict[tuple[int, int, int], dict] = {}
    sectores_por_modulo: dict[int, set[str]] = defaultdict(set)

    for ws in wb.worksheets:
        for fila in ws.iter_rows(values_only=True):
            for valor in fila:
                if not isinstance(valor, str):
                    continue

                texto = valor.strip()
                m = PATRON_MTV.match(texto)
                if m:
                    modulo = int(m.group(1))
                    turno = int(m.group(2))
                    valvula = int(m.group(3))
                    clave = (modulo, turno, valvula)
                    registros.setdefault(clave, {
                        'Fundo': fundo_default,
                        'Sector': None,
                        'Modulo': modulo,
                        'Turno': turno,
                        'Valvula': valvula,
                        'Cama': None,
                        'Codigo_SAP_Campo': None,
                        'Es_Test_Block': 0,
                    })
                    continue

                m = PATRON_SMTV.match(texto)
                if m:
                    sector = int(m.group(1))
                    modulo = int(m.group(2))
                    turno = int(m.group(3))
                    valvula = int(m.group(4))
                    clave = (modulo, turno, valvula)
                    sectores_por_modulo[modulo].add(f'Sector {sector}')
                    registros.setdefault(clave, {
                        'Fundo': fundo_default,
                        'Sector': f'Sector {sector}',
                        'Modulo': modulo,
                        'Turno': turno,
                        'Valvula': valvula,
                        'Cama': None,
                        'Codigo_SAP_Campo': None,
                        'Es_Test_Block': 0,
                    })

    for registro in registros.values():
        modulo = registro['Modulo']
        if registro['Sector'] is None and len(sectores_por_modulo[modulo]) == 1:
            registro['Sector'] = next(iter(sectores_por_modulo[modulo]))
        if registro['Sector'] is None and sector_default:
            registro['Sector'] = sector_default

    return sorted(
        registros.values(),
        key=lambda r: (r['Modulo'], r['Turno'], r['Valvula'])
    )


def guardar_csv(registros: list[dict], ruta_salida: Path) -> None:
    campos = [
        'Fundo',
        'Sector',
        'Modulo',
        'Turno',
        'Valvula',
        'Cama',
        'Codigo_SAP_Campo',
        'Es_Test_Block',
    ]
    with ruta_salida.open('w', newline='', encoding='utf-8-sig') as archivo:
        escritor = csv.DictWriter(archivo, fieldnames=campos)
        escritor.writeheader()
        escritor.writerows(registros)


def _sql_valor(valor) -> str:
    if valor is None:
        return 'NULL'
    if isinstance(valor, str):
        return "N'" + valor.replace("'", "''") + "'"
    return str(valor)


def guardar_sql(registros: list[dict], ruta_salida: Path) -> None:
    lineas = [
        'SET NOCOUNT ON;',
        '',
        'INSERT INTO MDM.Catalogo_Geografia (',
        '    Fundo,',
        '    Sector,',
        '    Modulo,',
        '    Turno,',
        '    Valvula,',
        '    Cama,',
        '    Codigo_SAP_Campo,',
        '    Es_Test_Block,',
        '    Es_Activa',
        ')',
        'SELECT *',
        'FROM (VALUES',
    ]

    valores = []
    for registro in registros:
        valores.append(
            '    ({}, {}, {}, {}, {}, {}, {}, {}, 1)'.format(
                _sql_valor(registro['Fundo']),
                _sql_valor(registro['Sector']),
                _sql_valor(registro['Modulo']),
                _sql_valor(registro['Turno']),
                _sql_valor(registro['Valvula']),
                _sql_valor(registro['Cama']),
                _sql_valor(registro['Codigo_SAP_Campo']),
                _sql_valor(registro['Es_Test_Block']),
            )
        )

    if valores:
        valores[-1] = valores[-1].rstrip(',')

    lineas.extend(',\n'.join(valores).splitlines())
    lineas.extend([
        ') AS src (',
        '    Fundo,',
        '    Sector,',
        '    Modulo,',
        '    Turno,',
        '    Valvula,',
        '    Cama,',
        '    Codigo_SAP_Campo,',
        '    Es_Test_Block,',
        '    Es_Activa',
        ')',
        'WHERE NOT EXISTS (',
        '    SELECT 1',
        '    FROM MDM.Catalogo_Geografia t',
        "    WHERE COALESCE(t.Fundo, N'')  = COALESCE(src.Fundo, N'')",
        "      AND COALESCE(t.Sector, N'') = COALESCE(src.Sector, N'')",
        '      AND t.Modulo  = src.Modulo',
        '      AND t.Turno   = src.Turno',
        '      AND t.Valvula = src.Valvula',
        ');',
        '',
    ])

    ruta_salida.write_text('\n'.join(lineas), encoding='utf-8')


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('--archivo', required=True, help='Ruta del plano en Excel')
    parser.add_argument('--fundo', default='ARANDANO ACP', help='Fundo canónico para los registros extraídos')
    parser.add_argument('--sector-default', default='SIN_SECTOR_MAPA', help='Sector por defecto si el plano no lo explicita')
    parser.add_argument('--salida-csv', default='geografia_plano_extraida.csv')
    parser.add_argument('--salida-sql', default='seed_mdm_catalogo_geografia.sql')
    args = parser.parse_args()

    ruta_excel = Path(args.archivo)
    registros = extraer_registros(
        ruta_excel,
        fundo_default=args.fundo,
        sector_default=args.sector_default,
    )

    guardar_csv(registros, Path(args.salida_csv))
    guardar_sql(registros, Path(args.salida_sql))

    modulos = sorted({r['Modulo'] for r in registros})
    print(f'Registros extraidos: {len(registros)}')
    print(f'Modulos detectados: {modulos}')
    print(f'CSV generado: {Path(args.salida_csv).resolve()}')
    print(f'SQL generado: {Path(args.salida_sql).resolve()}')


if __name__ == '__main__':
    main()
